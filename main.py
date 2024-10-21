import os
import re
from typing import List, Dict, Tuple, Optional, Any

import logzero
from bs4 import BeautifulSoup
import requests
from logzero import logger
from tqdm import tqdm
from multiprocessing import Pool

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SITEMAP_URL = "https://bikez.com/sitemap/motorcycle-specs.xml"
URL_LIST_FILENAME = "url.list"

N_PROCESS = 24

def parse_sitemap(sitemap: str) -> List[str]:
    logger.info(f"Retrieving sitemap URLs...")

    if os.path.isfile(URL_LIST_FILENAME):
        # Load URL list from cache
        with open(URL_LIST_FILENAME, "r") as f:
            url_list = f.read().splitlines()
            logger.info(f"Retrieved {len(url_list)} URLs from cache.")
    else:
        # Get URL list and save it to the cache
        r = requests.get(sitemap)
        soup = BeautifulSoup(r.text, "xml")
        url_list = [url.loc.string for url in soup.urlset.find_all("url")]
        with open(URL_LIST_FILENAME, "w+") as f:
            f.write('\n'.join(url_list))
            logger.info(f"Retrieved and cached {len(url_list)} URLs.")

    return url_list


def retrieve_entry(soup: BeautifulSoup, key: Any) -> Optional[str]:
    found = soup.find(["a", "b"], string=key)
    if found is None:
        return None

    # Special case for table entries that are also links
    if found.name == "a":
        return found.parent.parent.parent.find_all('td')[1].text

    return found.parent.parent.find_all('td')[1].text


def parse_power(soup: BeautifulSoup) -> Tuple[Optional[float], Optional[float]]:
    power_entry = retrieve_entry(soup, re.compile(R"Power\s{2,}|Power output|Output|Effect"))
    if power_entry is None:
        return None, None

    regex = R"(.+) HP \((.+)  kW\)\)"
    match = re.search(regex, power_entry)
    if match is None:
        logger.error(f"Cannot regex power: {power_entry}")
        return None, None

    return float(match[1].replace(',', '')), float(match[2].replace(',', ''))


def parse_weight(soup: BeautifulSoup, key: Any) -> Tuple[Optional[float], Optional[float]]:
    weight_entry = retrieve_entry(soup, key)
    if weight_entry is None:
        return None, None

    regex = R"(.+) kg \((.+) pounds\)"
    match = re.search(regex, weight_entry)
    if match is None:
        logger.error(f"Cannot regex weight: {weight_entry}")
        return None, None

    return float(match[1].replace(',', '')), float(match[2].replace(',', ''))


def parse_power_weight_ratio(soup: BeautifulSoup,
                             power_hp: Optional[float],
                             wet_weight_kg: Optional[float],
                             dry_weight_kg: Optional[float]) -> Optional[float]:
    p_w_r_entry = retrieve_entry(soup, re.compile(R"Power/weight"))
    if p_w_r_entry is None:
        if power_hp:
            if wet_weight_kg:
                return power_hp / wet_weight_kg
            elif dry_weight_kg:
                return power_hp / dry_weight_kg

        return None

    regex = R"(.+) HP/kg"
    match = re.search(regex, p_w_r_entry)
    if match is None:
        logger.error(f"Cannot regex p/w ratio: {p_w_r_entry}")
        return None

    return float(match[1].replace(',', ''))


def retrieve_page(url: str):
    logzero.loglevel(logzero.CRITICAL)
    r = requests.get(url)
    soup = BeautifulSoup(r.text, "html.parser")

    name = retrieve_entry(soup, re.compile(R"Model|Motorcycle name"))
    logger.info(name)
    power_hp, power_kw = parse_power(soup)
    wet_weight_kg, wet_weight_lb = parse_weight(soup, re.compile(R"Weight incl. oil"))
    dry_weight_kg, dry_weight_lb = parse_weight(soup, re.compile(R"Dry weight"))
    year = retrieve_entry(soup, re.compile(R"\s{2,}Year|Year model|Year of manufacture|Model year"))
    if year is not None:
        year = int(year)

    entries = {"model": name,
               "year": year,
               "power_hp": power_hp,
               "power_kw": power_kw,
               "torque": retrieve_entry(soup, "Torque"),
               "displacement": retrieve_entry(soup, "Displacement"),
               "wet_weight_kg": wet_weight_kg,
               "wet_weight_lb": wet_weight_lb,
               "dry_weight_kg": dry_weight_kg,
               "dry_weight_lb": dry_weight_lb,
               "power_weight_ratio_hp_kg": parse_power_weight_ratio(soup, power_hp, wet_weight_kg, dry_weight_kg),
               "url": url}
    return entries


def retrieve_data(url_list: List[str]) -> List[Dict[str, str]]:
    logger.debug("Starting the pool.")
    with Pool(N_PROCESS) as pool:
        results = []
        for result in tqdm(pool.imap_unordered(retrieve_page, url_list), total=len(url_list)):
            results.append(result)
        return results


def export_data(data: List[Dict[str, str]]) -> None:
    logger.info("Exporting data as Excel spreadsheet")
    wb = Workbook()
    ws = wb.active

    ws.append(list(data[0].keys()))
    for row in data:
        ws.append(list(row.values()))

    tab = Table(displayName="Data", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                                        showFirstColumn=False,
                                        showLastColumn=False,
                                        showRowStripes=True,
                                        showColumnStripes=False)
    ws.add_table(tab)

    col_width = [91,   # model
                 7,    # year
                 12,   # power_hp
                 12,   # power_kw
                 45,   # torque
                 29,   # displacement
                 17,   # wet_weight_kg
                 16,   # wet_weight_lb
                 16,   # dry_weight_kg
                 16,   # dry_weight_lb
                 28,   # power_weight_ratio_hp_kg
                 131]  # url

    for i, column_cells in enumerate(ws.columns):
        ws.column_dimensions[column_cells[0].column_letter].width = col_width[i]

    wb.save("jerk.xlsx")


if __name__ == "__main__":
    logzero.loglevel(logzero.CRITICAL)
    urls = parse_sitemap(SITEMAP_URL)
    data = retrieve_data(urls)
    export_data(data)
